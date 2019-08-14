import re
from openpyxl import Workbook
import time
def getString(filename):
    with open(filename,'r') as f:
        text=f.read()
    return text

def parseRedString(str):
    pattern=re.compile('<TR>.*?bgColor=3D#ff8181.*?\.htm">(.*?)</A>.*?color=3Dblue>(.*?)</FONT>.*?\.write\(\'(.*?)\'\).*?</FONT></TD></TR>',re.S)
    items=re.findall(pattern,str)
    return items

def parseYellowString(str):
    pattern=re.compile('<TR>.*?bgColor=3Dyellow.*?<CENTER>(.*?)</CENTER>.*?3Dblue>(.*?)</FONT>.*?write\(\'(.*?)\'\).*?</FONT></TD></TR>',re.S)
    items=re.findall(pattern,str)
    return items

def data_write(items):
    wb=Workbook() #创建文件对象
    ws=wb.active #获取第一个sheet
    i=1
    for item in items:
        #print(item)
        ws["A"+str(i)]=item[0]
        ws["B"+str(i)]=item[1]
        ws["C"+str(i)]=item[2]
        i+=1
    wb.save("d:\\test.xlsx")  #保存

if __name__ == '__main__':
    text=getString("test.mht")
    redItems=parseRedString(text)
    yellowItems=parseYellowString(text)
    data_write(redItems+yellowItems)
    print("complete")

