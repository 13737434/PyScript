import re

from openpyxl import load_workbook

from 圈复杂度脚本.old.CyclomaticItem import CyclomaticItem
class Run:
    def __init__(self):
        self.cyclomaticItem=CyclomaticItem()
        self.items=[]

    #获取 XXXXX.txt 中的文本
    def getString(self, url):
        with open(url, 'r') as f:
            text=f.read()
        return text

    #正则提取所有函数的Cyclomatic     exp:('app_inf_get_sdpbtm_par', '4')
    def getAllCyclomatic(self,str):
        #\s 匹配任何空白字符，包括空格、制表符、换页符等等。等价于[ \f\n\r\t\v]。
        #\S 匹配任何非空白字符。等价于[^ \f\n\r\t\v]。
        pattern=re.compile('\s*(\S*)\s*Cyclomatic:\s*(\d+)\s*Modified\s*Cyclomatic:',re.S)
        items=re.findall(pattern,str)
        d={}
        for item in items:
            d[item[0]]=item[1]
        return d

    #正则匹配所有函数所在的.c文件 返回一个字典
    def getAllFunDict(self,str):
        pattern=re.compile('\s*(\S*)\s*\(Function\)\s*\[(.*?\.c),',re.S)
        items=re.findall(pattern,str)
        d={}
        for item in items:
            d[item[0]]=item[1]
        return d

    #写入 Excel 文档
    def data_write(self,dic_fileName,dic_cyclomatic):
        wb=load_workbook('Demo.xlsx')  #获取Excel文件对象
        ws=wb.active #获取第一个sheet
        i=12
        for item in self.items:
            if int(dic_cyclomatic[item])>20:
                ws["A"+str(i)]=i-11
                ws["B"+str(i)]=dic_fileName[item]
                ws["C"+str(i)]=item
                ws["D"+str(i)]=self.cyclomaticItem.serious_lv
                ws["E"+str(i)]=self.cyclomaticItem.find_sub
                ws["F"+str(i)]=self.cyclomaticItem.NCR_class
                ws["G"+str(i)]=self.cyclomaticItem.find_time
                ws["H"+str(i)]=self.cyclomaticItem.find_person
                ws["I"+str(i)]="圈复杂度为"+dic_cyclomatic[item]+"，大于20"
                ws["J"+str(i)]=self.cyclomaticItem.question_num
                ws["K"+str(i)]=self.cyclomaticItem.deal_opinion
                ws["L"+str(i)]=self.cyclomaticItem.solution
                ws["M"+str(i)]=self.cyclomaticItem.confirm_person
                ws["N"+str(i)]=self.cyclomaticItem.verify_time
                ws["O"+str(i)]=self.cyclomaticItem.state
                i+=1
        wb.save("C:\\Users\\v5682\\Desktop\\静态度量问题报告单.xlsx")  #保存

#程序入口
    def run(self,url):
        text=self.getString(url)
        dic_fileName=self.getAllFunDict(text)
        dic_cyclomatic=self.getAllCyclomatic(text)
        self.data_write(dic_fileName,dic_cyclomatic)


if __name__ == '__main__':
    a=Run()
    #此处 添加/修改 需要查找的函数
    a.items=['app_inf_send_can','app_task_time_monitor','appl_inf_get_plat_can_data']
    #此处 设置 操作的txt文本路径
    a.run("Cyclomatic.txt")
    print("finish")
