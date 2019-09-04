import re
import openpyxl
import time
import os
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill,colors
class Deal:
    #获取 XXXXX.txt 中的文本
    def getString(self, url):
        with open(url, 'r') as f:
            text=f.read()
        return text

    # 判断文本是否符合要求
    def isTrue(self,text):
        pattern=re.compile('Program\sUnit\sComplexity',re.S)
        items=re.findall(pattern,text)
        if(len(items)>0):
            return True
        else:
            return False



    #正则提取所有函数的Cyclomatic  返回一个字典    exp:('app_inf_get_sdpbtm_par', '4')
    def getAllCyclomatic(self,text):
        #\s 匹配任何空白字符，包括空格、制表符、换页符等等。等价于[ \f\n\r\t\v]。
        #\S 匹配任何非空白字符。等价于[^ \f\n\r\t\v]。
        pattern=re.compile('\s(\w*)\s*Cyclomatic:\s*(\d+)\s*Modified\sCyclomatic:',re.S)
        items=re.findall(pattern,text)
        d={}
        for item in items:
            d[item[0]]=item[1]
        return d



    #正则匹配所有函数所在的.c文件 返回一个字典
    def getAllFunDict(self,text):
        pattern=re.compile('\s(\w*)\s*\(Function\)\s*\[(.*?\.c),',re.S)
        items=re.findall(pattern,text)
        d={}
        for item in items:
            d[item[0]]=item[1]
        return d
        


    #Excel操作
    def makeExcel(self,model,items,savepath):
        wb=openpyxl.Workbook()#创建Excel文件对象
        ws=wb.active #获取第一个sheet
        '''设置列宽'''
        ws.column_dimensions['A'].width = 8 #序号
        ws.column_dimensions['B'].width = 19 #文件名
        ws.column_dimensions['C'].width = 29 #函数名
        ws.column_dimensions['D'].width = 9 #严重程度
        ws.column_dimensions['E'].width = 12 #发现方式
        ws.column_dimensions['F'].width = 11 #NCR分类
        ws.column_dimensions['G'].width = 13 #发现时间
        ws.column_dimensions['H'].width = 9 #发现人
        ws.column_dimensions['I'].width = 37 #详细问题描述
        ws.column_dimensions['J'].width = 9 #问题个数
        ws.column_dimensions['K'].width = 9 #处理意见
        ws.column_dimensions['L'].width = 15 #解决措施
        ws.column_dimensions['M'].width = 9 #研发确认
        ws.column_dimensions['N'].width = 11 #验证时间
        ws.column_dimensions['O'].width = 11 #状态

        '''设置提示'''
        ws.merge_cells('C2:G3')
        ws['C2']='问题严重程度：致命、严重、一般、轻微、建议'
        ws['C2'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
        ws['C2'].font=Font(name='宋体', size=11)
        ws['C2'].fill=PatternFill(fill_type='solid',fgColor=colors.YELLOW)

        ws.merge_cells('C4:G5')
        ws['C4']='NCR分类：文档缺陷/编码缺陷'
        ws['C4'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
        ws['C4'].font=Font(name='宋体', size=11)
        ws['C4'].fill=PatternFill(fill_type='solid',fgColor=colors.YELLOW)

        ws.merge_cells('C6:G6')
        ws['C6'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
        ws['C6'].font=Font(name='宋体', size=11)
        ws['C6'].fill=PatternFill(fill_type='solid',fgColor=colors.YELLOW)

        ws.merge_cells('I2:K3')
        ws['I2']='填写说明：处理意见填写“说明”或者“修改”'
        ws['I2'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
        ws['I2'].font=Font(name='宋体', size=11)
        ws['I2'].fill=PatternFill(fill_type='solid',fgColor=colors.RED)

        ws.merge_cells('I4:K5')
        ws['I4']='若处理意见为说明，则需要在解决措施中填写该问题不修改的原因；'
        ws['I4'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
        ws['I4'].font=Font(name='宋体', size=11)
        ws['I4'].fill=PatternFill(fill_type='solid',fgColor=colors.RED)

        ws.merge_cells('I6:K7')
        ws['I6']='若处理意见为修改，则需要在解决措施中填写如何修改，并写明修改内容。'
        ws['I6'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
        ws['I6'].font=Font(name='宋体', size=11)
        ws['I6'].fill=PatternFill(fill_type='solid',fgColor=colors.RED)

        ws.merge_cells('I8:K8')
        ws['I8']='研发确认：处理问题的人员。'
        ws['I8'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
        ws['I8'].font=Font(name='宋体', size=11)
        ws['I8'].fill=PatternFill(fill_type='solid',fgColor=colors.RED)

        '''设置列标题'''
        ws["A11"]='序号'
        ws["B11"]='文件名'
        ws["C11"]='函数名'
        ws["D11"]='严重程度'
        ws["E11"]='发现方式'
        ws["F11"]='NCR分类'
        ws["G11"]='发现时间'
        ws["H11"]='发现人'
        ws["I11"]='详细问题描述'
        ws["J11"]='问题个数'
        ws["K11"]='处理意见'
        ws["L11"]='解决措施'
        ws["M11"]='研发确认'
        ws["N11"]='验证时间'
        ws["O11"]='状态'

        '''设置单元格格式'''
        for nn in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']:
            ws[nn+'11'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
            ws[nn+'11'].font=Font(name='宋体', size=11)
            border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin',color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin',color='000000'))
            ws[nn+'11'].border=border

        i=12
        for item in items:
            print(item)
            ws["A"+str(i)]=i-11
            ws["B"+str(i)]=item[2]
            ws["C"+str(i)]=item[0]
            ws["D"+str(i)]=model.seriouslv
            ws["E"+str(i)]=model.findsub
            ws["F"+str(i)]=model.NCRclass
            ws["G"+str(i)]=model.findtime
            ws["H"+str(i)]=model.findperson
            ws["I"+str(i)]="圈复杂度为"+item[1]+"，大于20"
            ws["J"+str(i)]=model.questionnum
            ws["K"+str(i)]=model.dealopinion
            ws["L"+str(i)]=model.solution
            ws["M"+str(i)]=model.confirmperson
            ws["N"+str(i)]=model.verifytime
            ws["O"+str(i)]=model.stated
            '''设置单元格格式'''
            for nn in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']:
                ws[nn+str(i)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
                ws[nn+str(i)].font=Font(name='宋体', size=11)
                border = Border(left=Side(border_style='thin',color='000000'),
                                right=Side(border_style='thin',color='000000'),
                                top=Side(border_style='thin',color='000000'),
                                bottom=Side(border_style='thin',color='000000'))
                ws[nn+str(i)].border=border
            i+=1

        wb.save(savepath+'\\静态度量问题报告单.xlsx')






if __name__ == '__main__':
    dd=Deal()
    text=dd.getString('C:/Users/v5682/Desktop/MyUnderstandProject.txt')
    # print(dd.getAllFunDict(text))
    print(dd.isTrue(text))







