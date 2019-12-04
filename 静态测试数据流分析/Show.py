import tkinter
import tkinter.filedialog
import tkinter.messagebox
import datetime
import os
import threading
import base64
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side
from 静态测试数据流分析.icon import img
from 静态测试数据流分析.Deal import Deal

class Show:
    #成员属性
    ischoose=False
    filelist=[]

    def __init__(self):
        self.nowtime=datetime.datetime.now().strftime('%Y%m%d') #当前时间
        self.UIInit()  # 界面初始化


    def UIInit(self):
        tmp = open("tmp.ico","wb+")
        tmp.write(base64.b64decode(img))
        tmp.close()
        self.root=tkinter.Tk()
        self.root.title('静态测试数据流分析-Script')
        self.root.geometry('500x300+100+200')
        self.root.resizable(0,0)
        self.root.iconbitmap('tmp.ico')
        os.remove("tmp.ico")

        # 不合格项等级
        self.txtUnqualifiedlv=tkinter.StringVar()
        self.txtUnqualifiedlv.set('建议')
        self.labUnqualifiedlv=tkinter.Label(self.root,text='不合格项等级:')
        self.labUnqualifiedlv.place(x=10,y=10,width=100,height=20)
        self.entUnqualifiedlv=tkinter.Entry(self.root,textvariable=self.txtUnqualifiedlv)
        self.entUnqualifiedlv.place(x=120,y=10,width=100,height=20)

        # 验证时间
        self.txtVerifyTime=tkinter.StringVar()
        self.txtVerifyTime.set(self.nowtime)
        self.labVerifyTime=tkinter.Label(self.root,text='验证时间  :')
        self.labVerifyTime.place(x=10,y=40,width=100,height=20)
        self.entVerifyTime=tkinter.Entry(self.root,textvariable=self.txtVerifyTime)
        self.entVerifyTime.place(x=120,y=40,width=100,height=20)

        # 处理方式
        self.txtDealWay=tkinter.StringVar()
        self.txtDealWay.set('说明')
        self.labDealWay=tkinter.Label(self.root,text='处理方式  :')
        self.labDealWay.place(x=10,y=70,width=100,height=20)
        self.entDealWay=tkinter.Entry(self.root,textvariable=self.txtDealWay)
        self.entDealWay.place(x=120,y=70,width=100,height=20)

        # 研发说明内容
        self.txtExplanation=tkinter.StringVar()
        self.txtExplanation.set('经分析，实现逻辑正确，符合设计')
        self.labExplanation=tkinter.Label(self.root,text='研发说明内容  :')
        self.labExplanation.place(x=10,y=100,width=100,height=20)
        self.entExplanation=tkinter.Entry(self.root,textvariable=self.txtExplanation)
        self.entExplanation.place(x=120,y=100,width=100,height=20)

        # 研发确认
        self.txtConfirmPerson=tkinter.StringVar()
        self.txtConfirmPerson.set('XXX')
        self.labConfirmPerson=tkinter.Label(self.root,text='研发确认  :')
        self.labConfirmPerson.place(x=10,y=130,width=100,height=20)
        self.entConfirmPerson=tkinter.Entry(self.root,textvariable=self.txtConfirmPerson)
        self.entConfirmPerson.place(x=120,y=130,width=100,height=20)

        # 测试者
        self.txtTestPerson=tkinter.StringVar()
        self.txtTestPerson.set('XXX')
        self.labTestPerson=tkinter.Label(self.root,text='测试者  :')
        self.labTestPerson.place(x=10,y=160,width=100,height=20)
        self.entTestPerson=tkinter.Entry(self.root,textvariable=self.txtTestPerson)
        self.entTestPerson.place(x=120,y=160,width=100,height=20)

        # 状态
        self.txtStated=tkinter.StringVar()
        self.txtStated.set('closed')
        self.labStated=tkinter.Label(self.root,text='状态  :')
        self.labStated.place(x=10,y=190,width=100,height=20)
        self.entStated=tkinter.Entry(self.root,textvariable=self.txtStated)
        self.entStated.place(x=120,y=190,width=100,height=20)

        # 文件/文件夹模式选择
        self.int_val=tkinter.IntVar()
        self.int_val.set('11')
        self.rbtnChooseFile=tkinter.Radiobutton(self.root,text='选择文件',variable=self.int_val,value="11")
        self.rbtnChooseFile.place(x=10,y=220,height=20)
        self.rbtnChooseFiles=tkinter.Radiobutton(self.root,text='选择文件夹',variable=self.int_val,value="22")
        self.rbtnChooseFiles.place(x=10,y=250,height=20)

        # 消息提示区
        self.labMessTitle=tkinter.Label(self.root,text='消息提示')
        self.labMessTitle.place(x=330,y=10,height=20)
        self.txtMessShow=tkinter.Text(self.root, bg="#ffffff")
        self.txtMessShow.place(x=240, y=30, width=250, height=150)
        self.scroll = tkinter.Scrollbar(self.txtMessShow)
        self.scroll.pack(side=tkinter.RIGHT,fill=tkinter.Y)
        self.txtMessShow.config(yscrollcommand=self.scroll.set)

        # 添加按钮
        self.btnAdd=tkinter.Button(self.root,text='添加',command=self.clickAdd)
        self.btnAdd.place(x=120,y=230,width=90,height=30)

        # 导出按钮
        self.btnExport=tkinter.Button(self.root,text='生成',command=self.clickMaker)
        self.btnExport.place(x=230,y=230,width=90,height=30)

        # 帮助
        self.labhelp=tkinter.Label(self.root,text='帮助',font = ('宋体', 10),fg='#0000ff')
        self.labhelp.place(x=360,y=210,width=90,height=30)
        self.labhelp.bind('<Button-1>',self.clickHelp)


        self.root.mainloop()

    # 帮助按钮点击事件
    def clickHelp(self,e):
        self.txtMessShow.insert(tkinter.END, '错误反馈及建议 379221379@qq.com\n--------------------------------\n')
        self.txtMessShow.see(tkinter.END)


    # 添加按钮点击事件
    def clickAdd(self):
        self.txtMessShow.delete('1.0','end')
        if(self.int_val.get()==11):
            self.filelist=tkinter.filedialog.askopenfilenames(title='请选择数据流报告',filetypes=(("HTML files", "*.mht;"),))
            self.txtMessShow.insert(tkinter.END, '选择了' + str(len(self.filelist)) + '个文件\n--------------------------------\n')
            self.txtMessShow.see(tkinter.END)
            if(len(self.filelist)>0):
                self.ischoose=True
            else:
                self.ischoose=False
        elif(self.int_val.get()==22):
            self.path=tkinter.filedialog.askdirectory(title='请选择文件夹')  #可遍历整个文件夹中所有的.mht文件
            if(self.path):
                self.filelist=self.selectFile(self.path)
                self.txtMessShow.insert(tkinter.END, '选择了' + str(len(self.filelist)) + '个文件\n--------------------------------\n')
                self.txtMessShow.see(tkinter.END)
                if(len(self.filelist)>0):
                    self.ischoose=True
                else:
                    self.ischoose=False
            else:
                self.txtMessShow.insert(tkinter.END, '未选择文件夹\n--------------------------------\n')
                self.txtMessShow.see(tkinter.END)

        else:
            self.txtMessShow.insert(tkinter.END, 'error\n')
            self.txtMessShow.see(tkinter.END)

    # 生成按钮点击事件
    def clickMaker(self):
        if(self.ischoose):
            self.savepath=tkinter.filedialog.askdirectory(title='请选择保存路径')
            if(self.savepath):
                def thread1(filelist):
                    try:
                        dd=Deal()
                        self.errs=""
                        self.txtMessShow.insert(tkinter.END, '正在生成.\n--------------------------------\n')
                        self.txtMessShow.see(tkinter.END)
                        tmp=1
                        for file in filelist:
                            self.errs=file
                            dd.getString(file)
                            dd.getBS()
                            if(dd.txtIsTrue()):
                                cname=dd.getCName()
                                if(cname=='unknow'):
                                    self.dataWrite(cname+str(tmp),dd.getAllTable(),self.savepath)
                                    self.txtMessShow.insert(tkinter.END, '请修改unknow文件.\n--------------------------------\n')
                                    self.txtMessShow.see(tkinter.END)
                                    tmp+=1
                                else:
                                    self.dataWrite(cname,dd.getAllTable(),self.savepath)
                                self.txtMessShow.insert(tkinter.END,cname+'----导出完成\n--------------------------------\n')
                                self.txtMessShow.see(tkinter.END)
                        self.txtMessShow.insert(tkinter.END, '任务结束.\n--------------------------------\n')
                        self.txtMessShow.see(tkinter.END)
                        tkinter.messagebox.showinfo("Finish","任务结束.")
                        self.ischoose=False
                    except Exception as exc:
                        tkinter.messagebox.showerror("Finish","请检查"+str(self.errs)+"\n"+str(exc))
                        print(exc)
                th=threading.Thread(target=thread1,args=(self.filelist,))
                th.setDaemon(True)
                th.start()
            else:
                self.txtMessShow.insert(tkinter.END, '未选择保存路径\n--------------------------------\n')
                self.txtMessShow.see(tkinter.END)
        else:
            self.txtMessShow.insert(tkinter.END, '没有可操作的文件\n--------------------------------\n')
            self.txtMessShow.see(tkinter.END)

    #递归遍历筛选文件
    def selectFile(self,dirpath):
        filelist=[]
        for root, dirs, files in os.walk(dirpath):
            #print(root) #当前目录路径
            #print(dirs) #当前路径下所有子目录
            #print(files) #当前路径下所有非目录子文件.
            for file in files:
                if os.path.splitext(file)[1] == '.mht':
                    filelist.append(os.path.join(root, file))
        return filelist

    #绘制EXCEL
    def dataWrite(self,cname,tables,savepath):
        txtUnqualifiedlv=self.txtUnqualifiedlv.get()
        txtVerifyTime=self.txtVerifyTime.get()
        txtDealWay=self.txtDealWay.get()
        txtExplanation=self.txtExplanation.get()
        txtConfirmPerson=self.txtConfirmPerson.get()
        txtTestPerson=self.txtTestPerson.get()
        txtStated=self.txtStated.get()
        wb=openpyxl.Workbook()#创建Excel文件对象

        #UR
        ws=wb.active #获取第一个sheet
        ws.title='UR'
        ws["A1"]='序号'
        ws["B1"]='文件名'
        ws["C1"]='Variable'
        ws["D1"]='Undefine'
        ws["E1"]='Reference'
        ws["F1"]='详细'
        ws["G1"]='不合格项等级'
        ws["H1"]='验证时间'
        ws["I1"]='处理方式'
        ws["J1"]='研发说明内容'
        ws["K1"]='研发确认'
        ws["L1"]='测试者'
        ws["M1"]='状态'
        '''设置列宽'''
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 26
        ws.column_dimensions['C'].width = 53
        ws.column_dimensions['D'].width = 9
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 59
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 8
        ws.column_dimensions['J'].width = 19
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 8
        ws.column_dimensions['M'].width = 8

        '''设置单元格格式'''
        for nn in ['A','B','C','D','E','F','G','H','I','J','K','L','M']:
            ws[nn+'1'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
            ws[nn+'1'].font=Font(name='宋体', size=11)
            border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin',color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin',color='000000'))
            ws[nn+'1'].border=border
        i=2

        for item in tables[0]:
            ws["A"+str(i)]=i-1
            ws["B"+str(i)]=cname
            ws["C"+str(i)]=item[0]
            ws["D"+str(i)]=item[1]
            ws["E"+str(i)]=item[2]
            ws["F"+str(i)]=item[3]
            ws["G"+str(i)]=txtUnqualifiedlv
            ws["H"+str(i)]=txtVerifyTime
            ws["I"+str(i)]=txtDealWay
            ws["J"+str(i)]=txtExplanation
            ws["K"+str(i)]=txtConfirmPerson
            ws["L"+str(i)]=txtTestPerson
            ws["M"+str(i)]=txtStated

            '''设置单元格格式'''
            for nn in ['A','B','C','D','E','F','G','H','I','J','K','L','M']:
                ws[nn+str(i)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
                ws[nn+str(i)].font=Font(name='宋体', size=11)
                border = Border(left=Side(border_style='thin',color='000000'),
                                right=Side(border_style='thin',color='000000'),
                                top=Side(border_style='thin',color='000000'),
                                bottom=Side(border_style='thin',color='000000'))
                ws[nn+str(i)].border=border
            i+=1

        #DU
        ws=wb.create_sheet(title='DU')
        ws["A1"]='序号'
        ws["B1"]='文件名'
        ws["C1"]='Variable'
        ws["D1"]='Define'
        ws["E1"]='Undefine'
        ws["F1"]='详细'
        ws["G1"]='不合格项等级'
        ws["H1"]='验证时间'
        ws["I1"]='处理方式'
        ws["J1"]='研发说明内容'
        ws["K1"]='研发确认'
        ws["L1"]='测试者'
        ws["M1"]='状态'
        '''设置列宽'''
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 26
        ws.column_dimensions['C'].width = 53
        ws.column_dimensions['D'].width = 9
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 59
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 8
        ws.column_dimensions['J'].width = 19
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 8
        ws.column_dimensions['M'].width = 8

        '''设置单元格格式'''
        for nn in ['A','B','C','D','E','F','G','H','I','J','K','L','M']:
            ws[nn+'1'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
            ws[nn+'1'].font=Font(name='宋体', size=11)
            border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin',color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin',color='000000'))
            ws[nn+'1'].border=border
        i=2

        for item in tables[1]:
            ws["A"+str(i)]=i-1
            ws["B"+str(i)]=cname
            ws["C"+str(i)]=item[0]
            ws["D"+str(i)]=item[1]
            ws["E"+str(i)]=item[2]
            ws["F"+str(i)]=item[3]
            ws["G"+str(i)]=txtUnqualifiedlv
            ws["H"+str(i)]=txtVerifyTime
            ws["I"+str(i)]=txtDealWay
            ws["J"+str(i)]=txtExplanation
            ws["K"+str(i)]=txtConfirmPerson
            ws["L"+str(i)]=txtTestPerson
            ws["M"+str(i)]=txtStated

            '''设置单元格格式'''
            for nn in ['A','B','C','D','E','F','G','H','I','J','K','L','M']:
                ws[nn+str(i)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
                ws[nn+str(i)].font=Font(name='宋体', size=11)
                border = Border(left=Side(border_style='thin',color='000000'),
                                right=Side(border_style='thin',color='000000'),
                                top=Side(border_style='thin',color='000000'),
                                bottom=Side(border_style='thin',color='000000'))
                ws[nn+str(i)].border=border
            i+=1

        #DD
        ws=wb.create_sheet(title='DD')
        ws["A1"]='序号'
        ws["B1"]='文件名'
        ws["C1"]='Variable'
        ws["D1"]='Define'
        ws["E1"]='Redefine'
        ws["F1"]='详细'
        ws["G1"]='不合格项等级'
        ws["H1"]='验证时间'
        ws["I1"]='处理方式'
        ws["J1"]='研发说明内容'
        ws["K1"]='研发确认'
        ws["L1"]='测试者'
        ws["M1"]='状态'
        '''设置列宽'''
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 26
        ws.column_dimensions['C'].width = 53
        ws.column_dimensions['D'].width = 9
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 59
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 8
        ws.column_dimensions['J'].width = 19
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 8
        ws.column_dimensions['M'].width = 8

        '''设置单元格格式'''
        for nn in ['A','B','C','D','E','F','G','H','I','J','K','L','M']:
            ws[nn+'1'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
            ws[nn+'1'].font=Font(name='宋体', size=11)
            border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin',color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin',color='000000'))
            ws[nn+'1'].border=border
        i=2

        for item in tables[2]:
            ws["A"+str(i)]=i-1
            ws["B"+str(i)]=cname
            ws["C"+str(i)]=item[0]
            ws["D"+str(i)]=item[1]
            ws["E"+str(i)]=item[2]
            ws["F"+str(i)]=item[3]
            ws["G"+str(i)]=txtUnqualifiedlv
            ws["H"+str(i)]=txtVerifyTime
            ws["I"+str(i)]=txtDealWay
            ws["J"+str(i)]=txtExplanation
            ws["K"+str(i)]=txtConfirmPerson
            ws["L"+str(i)]=txtTestPerson
            ws["M"+str(i)]=txtStated

            '''设置单元格格式'''
            for nn in ['A','B','C','D','E','F','G','H','I','J','K','L','M']:
                ws[nn+str(i)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
                ws[nn+str(i)].font=Font(name='宋体', size=11)
                border = Border(left=Side(border_style='thin',color='000000'),
                                right=Side(border_style='thin',color='000000'),
                                top=Side(border_style='thin',color='000000'),
                                bottom=Side(border_style='thin',color='000000'))
                ws[nn+str(i)].border=border
            i+=1

        wb.save(savepath+'\\'+cname+ "-静态分析数据流分析问题报告单.xlsx")  #保存


if __name__ == '__main__':
    Show()