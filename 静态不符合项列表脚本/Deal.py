import re
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side
class Deal:
    def __init__(self,Model):
        self.model=Model
        '''正则初始化'''
        self.patternfilename=re.compile('<H1>File.*?(\w+)\\.c.*?</H1>',re.S)
        self.patternred=re.compile('<TR>.*?bgColor=3D#ff8181.*?\.htm">(.*?)</A>.*?color=3Dblue>(.*?)</FONT>.*?\.write\(\'(.*?)\'\).*?</FONT></TD></TR>',re.S)
        self.patternyellow=re.compile('<TR>.*?bgColor=3Dyellow.*?<CENTER>(.*?)</CENTER>.*?3Dblue>(.*?)</FONT>.*?write\(\'(.*?)\'\).*?</FONT></TD></TR>',re.S)

    #获取 XXXXX.mht 中的文本
    def getString(self, url):
        with open(url, 'r') as f:
            text=f.read()
        return text

    #获取文件名
    def getFileName(self,str):
        file_name=re.findall(self.patternfilename,str)[0].replace("\n", "").replace("=", "")
        return file_name

    #解析红色部分
    def parseRedString(self,str):
        items=re.findall(self.patternred,str)
        return items
    #解析黄色部分
    def parseYellowString(self,str):
        items=re.findall(self.patternyellow,str)
        return items

    #写入 Excel 文档
    def dataWrite(self,items,savepath):
        wb=openpyxl.Workbook()#创建Excel文件对象
        #wb=load_workbook('Demo.xlsx')  #获取Excel文件对象
        ws=wb.active #获取第一个sheet
        ws["A1"]='序号'
        ws["B1"]='文件名'
        ws["C1"]='测试技术'
        ws["D1"]='不合格项'
        ws["E1"]='不合格项个数'
        ws["F1"]='不合格项等级'
        ws["G1"]='验证时间'
        ws["H1"]='处理方式'
        ws["I1"]='研发说明内容'
        ws["J1"]='研发确认'
        ws["K1"]='测试者'
        ws["L1"]='状态'
        '''设置列宽'''
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 17
        ws.column_dimensions['C'].width = 17
        ws.column_dimensions['D'].width = 64
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 19
        ws.column_dimensions['J'].width = 8
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 8

        '''设置单元格格式'''
        for nn in ['A','B','C','D','E','F','G','H','I','J','K','L']:
            ws[nn+'1'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
            ws[nn+'1'].font=Font(name='宋体', size=11)
            border = Border(left=Side(border_style='thin',color='000000'),
                        right=Side(border_style='thin',color='000000'),
                        top=Side(border_style='thin',color='000000'),
                        bottom=Side(border_style='thin',color='000000'))
            ws[nn+'1'].border=border

        i=2
        for item in items:
            ws["A"+str(i)]=i-1
            ws["B"+str(i)]= self.model.filename+ ".c"
            ws["C"+str(i)]=self.model.testskill
            ws["D"+str(i)]=item[1].replace("\n", "").replace("=", "")+':'+item[2].replace("\n", "").replace("=", "")
            ws["E"+str(i)]=item[0].replace("\n", "").replace("=", "")
            ws["F"+str(i)]=self.model.disqualificationlv
            ws["G"+str(i)]=self.model.verifytime
            ws["H"+str(i)]=self.model.dealway
            ws["I"+str(i)]=self.model.explanation
            ws["J"+str(i)]=self.model.confirmperson
            ws["K"+str(i)]=self.model.testperson
            ws["L"+str(i)]=self.model.stated

            '''设置单元格格式'''
            for nn in ['A','B','C','D','E','F','G','H','I','J','K','L']:
                ws[nn+str(i)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
                ws[nn+str(i)].font=Font(name='宋体', size=11)
                border = Border(left=Side(border_style='thin',color='000000'),
                                right=Side(border_style='thin',color='000000'),
                                top=Side(border_style='thin',color='000000'),
                                bottom=Side(border_style='thin',color='000000'))
                ws[nn+str(i)].border=border
            i+=1
        wb.save(savepath+'\\'+ self.model.filename + "-静态不符合项列表.xlsx")  #保存

    #生成一份报告
    def getOneReport(self,filepath,savepath):
        text=self.getString(filepath)
        self.model.filename=self.getFileName(text)
        print(self.model.filename)
        redItems=self.parseRedString(text)
        yellowItems=self.parseYellowString(text)
        self.dataWrite(redItems+yellowItems,savepath)
        return self.model.filename
    #生成所有的.mht文件的报告
    # def getAllReport(self,filelist):
    #     for file in filelist:
    #         self.getOneReport(file)



