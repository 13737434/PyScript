import re
from openpyxl import Workbook, load_workbook
from 静态不符合项列表脚本.TestItem import TheSameItem

class ProductOne:
    def __init__(self):
        self.theSameItem=TheSameItem()

    #获取 XXXXX.mht 中的文本
    def getString(self, url):
        with open(url, 'r') as f:
            text=f.read()
        return text

    #获取文件名
    def getFileName(self,str):
        pattern=re.compile(r'<H1>File.*\\(.*?)\.c.*?</H1>',re.S)
        file_name=re.findall(pattern,str)[0]
        return file_name

    #解析红色部分
    def parseRedString(self,str):
        pattern=re.compile('<TR>.*?bgColor=3D#ff8181.*?\.htm">(.*?)</A>.*?color=3Dblue>(.*?)</FONT>.*?\.write\(\'(.*?)\'\).*?</FONT></TD></TR>',re.S)
        items=re.findall(pattern,str)
        return items
    #解析蓝色部分
    def parseYellowString(self,str):
        pattern=re.compile('<TR>.*?bgColor=3Dyellow.*?<CENTER>(.*?)</CENTER>.*?3Dblue>(.*?)</FONT>.*?write\(\'(.*?)\'\).*?</FONT></TD></TR>',re.S)
        items=re.findall(pattern,str)
        return items

    #写入 Excel 文档
    def data_write(self,items):
        #wb=Workbook() #创建Excel文件对象
        wb=load_workbook('Demo.xlsx')  #获取Excel文件对象
        ws=wb.active #获取第一个sheet
        i=2
        for item in items:
            ws["A"+str(i)]=i-1
            ws["B"+str(i)]= self.theSameItem.file_name + ".c"
            ws["C"+str(i)]=self.theSameItem.test_skill
            ws["D"+str(i)]=item[1]+':'+item[2]
            ws["E"+str(i)]=item[0]
            ws["F"+str(i)]=self.theSameItem.disqualification_lv
            ws["G"+str(i)]=self.theSameItem.verify_time
            ws["H"+str(i)]=self.theSameItem.deal_way
            ws["I"+str(i)]=self.theSameItem.explanation
            ws["J"+str(i)]=self.theSameItem.confirm_person
            ws["K"+str(i)]=self.theSameItem.test_person
            ws["L"+str(i)]=self.theSameItem.state
            i+=1
        wb.save("C:\\Users\\v5682\\Desktop\\" + self.theSameItem.file_name + "-静态不符合项列表.xlsx")  #保存

    #生成一份报告
    def getOneReport(self,url):
        text=self.getString(url)
        self.theSameItem.file_name=self.getFileName(text)
        redItems=self.parseRedString(text)
        yellowItems=self.parseYellowString(text)
        self.data_write(redItems+yellowItems)


if __name__ == '__main__':
    a=ProductOne()
    a.getOneReport("C:\\Users\\v5682\\Desktop\\test.mht")