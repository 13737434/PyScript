import re
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill,colors
class Deal:
    # 构建正则表达式
    patternistrue=re.compile('Simple\s*Invocation\s*Tree',re.S)#解析文本是否符合要求
    patternpages=re.compile('Simple\s*Invocation\s*Tree\s*Report\s*=*\n(.*?)\x0c',re.S)#解析所有Simple Invocation Tree页内容
    patternfuns=re.compile(r'(\w+)\n(\|.*?)\n\n',re.S)#解析（函数+函数所调用函数）块
    patternfuncalls=re.compile('\w+',re.S)#解析每个（函数+函数所调用函数）块 中的函数，所调函数

    #获取 XXXXX.txt 中的文本
    def getString(self, url):
        with open(url, 'r') as f:
            text=f.read()
        return text

    # 判断文本是否符合要求
    def isTrue(self,text):
        items=re.findall(self.patternistrue,text)
        if(len(items)>0):
            return True
        else:
            return False

    #开始解析内容  调用level1：【调用level2，，，，，】
    def getFunDict(self,text):
        fundic={}
        pages=re.findall(self.patternpages,text)
        newtext=''.join(pages)+'\n'
        funlists=re.findall(self.patternfuns,newtext)
        for funli in funlists:
            calllists=re.findall(self.patternfuncalls,funli[1])
            fundic[funli[0]]=calllists
        return fundic

    #获得字典 调用level1：【入口函数名称，，，，，，】
    def getFunDict_1(self,fundic):
        fundic_1={}
        for f in fundic:
            tmplist=[]
            for li in fundic:
                if(f in fundic[li]):
                    tmplist.append(li)
            fundic_1[f]=tmplist
        return fundic_1

    #Excel操作 全部
    def makeExcel_all(self,fundic,fundic_1,savepath):
        wb=openpyxl.Workbook()#创建Excel文件对象
        ws=wb.active #获取第一个sheet
        '''设置列宽'''
        ws.column_dimensions['A'].width =10 #序号
        ws.column_dimensions['B'].width = 35 #入口函数名
        ws.column_dimensions['C'].width =35 #调用level1
        ws.column_dimensions['D'].width =35 #调用level2
        '''设置列标题'''
        ws["A1"]='序号'
        ws["B1"]='入口函数名'
        ws["C1"]='调用level1'
        ws["D1"]='调用level2'

        '''设置单元格格式'''
        for nn in ['A','B','C','D']:
            ws[nn+'1'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
            ws[nn+'1'].font=Font(name='宋体', size=11)
            border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin',color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin',color='000000'))
            ws[nn+'1'].border=border

        i=2
        for fd in fundic:
            ws["A"+str(i)]=i-1
            ws["B"+str(i)]='\n'.join(fundic_1[fd])
            ws["C"+str(i)]=fd
            ws["D"+str(i)]='\n'.join(fundic[fd])
            '''设置单元格格式'''
            for nn in ['A','B','C','D']:
                ws[nn+str(i)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
                ws[nn+str(i)].font=Font(name='宋体', size=11)
                border = Border(left=Side(border_style='thin',color='000000'),
                                right=Side(border_style='thin',color='000000'),
                                top=Side(border_style='thin',color='000000'),
                                bottom=Side(border_style='thin',color='000000'))
                ws[nn+str(i)].border=border
            i+=1

        wb.save(savepath+'\\函数调用关系(全部).xlsx')

    #Excel操作 部分
    def makeExcel(self,funlist,fundic,fundic_1,savepath):
        wb=openpyxl.Workbook()#创建Excel文件对象
        ws=wb.active #获取第一个sheet
        '''设置列宽'''
        ws.column_dimensions['A'].width =10 #序号
        ws.column_dimensions['B'].width = 35 #入口函数名
        ws.column_dimensions['C'].width =35 #调用level1
        ws.column_dimensions['D'].width =35 #调用level2
        '''设置列标题'''
        ws["A1"]='序号'
        ws["B1"]='入口函数名'
        ws["C1"]='调用level1'
        ws["D1"]='调用level2'

        '''设置单元格格式'''
        for nn in ['A','B','C','D']:
            ws[nn+'1'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
            ws[nn+'1'].font=Font(name='宋体', size=11)
            border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin',color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin',color='000000'))
            ws[nn+'1'].border=border

        i=2
        for fd in funlist:
            ws["A"+str(i)]=i-1
            ws["B"+str(i)]='\n'.join(fundic_1[fd])
            ws["C"+str(i)]=fd
            ws["D"+str(i)]='\n'.join(fundic[fd])
            '''设置单元格格式'''
            for nn in ['A','B','C','D']:
                ws[nn+str(i)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
                ws[nn+str(i)].font=Font(name='宋体', size=11)
                border = Border(left=Side(border_style='thin',color='000000'),
                                right=Side(border_style='thin',color='000000'),
                                top=Side(border_style='thin',color='000000'),
                                bottom=Side(border_style='thin',color='000000'))
                ws[nn+str(i)].border=border
            i+=1

        wb.save(savepath+'\\函数调用关系.xlsx')












if __name__ == '__main__':
    dd=Deal()
    text=dd.getString('C:/Users/v5682/Desktop/MyUnderstandProject.txt')
    #print(dd.isTrue(text))
    fundic=dd.getFunDict(text)
    fundic_1=dd.getFunDict_1(fundic)
    dd.makeExcel(fundic,fundic_1,"C:/Users/v5682/Desktop")


