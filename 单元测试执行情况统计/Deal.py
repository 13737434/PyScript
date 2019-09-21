import os
import re
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side
class Deal:
    resultlist={}
    casenumlist={}
    patterntype=re.compile(r'<TITLE>(.*?)</TITLE>',re.S) #解析mht文件类型

    '''Dynamic Coverage Analysis Report'''
    # patternfrom=re.compile(r'',re.S)#解析所在.c文件
    patternRows=re.compile(r'<TR>\s*<TD colSpan=3D11><A.*?<TD></TD></TR>',re.S) #解析Dynamic Coverage Analysis Report中某一函数的一行
    patternResult=re.compile('<TD colSpan=3D11><A.*?>(.*?)</A></TD>.*?>\+?([\d\[].*?)<.*?>\+?([\d\[].*?)<.*?>\+?([\d\[].*?)<',re.S) #解析覆盖率
    patternResultre=re.compile('<TD colSpan=3D11><A.*?>(.*?)</A></TD>.*?>\+?([\d\[].*?)<.*?>\+?([\d\[].*?)<',re.S)#解析失败，再次解析
    ''''''

    '''LDRA TBrun Regression Report'''
    patternRows_1=re.compile(r'<TR align=\w*?>\s*<TD><A.*?</TD></TR>',re.S) #解析某一行
    patternfrom=re.compile(r'<TD>(\w.*?)</TD>',re.S) #解析案例数、所在.c


    ''''''


    #获取 XXXXX.mht 中的文本
    def getText(self, url):
        with open(url, 'r') as f:
            text=f.read()
        return text

    #处理文本
    def dealText(self,text):
        texttype=re.findall(self.patterntype,text)[0]
        if(texttype=='LDRA Dynamic Coverage Analysis'):
            self.dealDynamic(text)
        elif(texttype=='LDRA TBrun Regression Report'):
            self.dealRegression(text)
        else:
            print('error')

    #Deal Dynamic Coverage Analysis Report
    def dealDynamic(self,text):
        funlist=re.findall(self.patternRows,text)
        for afun in funlist:
            r=re.findall(self.patternResult,afun)
            if(not r):#解析失败，再次解析 此次条件是没有MC/DC的情况
                r=re.findall(self.patternResultre,afun)
                if(r):
                    funname=r[0][0].replace("\n", "").replace("=20", "").replace("=", "").replace(" ", "")
                    statement=r[0][1].replace("\n", "").replace("=20", "").replace("=", "").replace(" ", "")
                    branch=r[0][2].replace("\n", "").replace("=20", "").replace("=", "").replace(" ", "")
                    mcdc='NULL'
                    if(int(statement)>0):
                        if(funname in self.resultlist): #在字典中，比较大小取最大
                            if(int(statement)>int(self.resultlist[funname][0])): #statment部分大于
                                if(len(branch)>3):
                                        self.resultlist[funname]=[statement,'NULL','NULL']
                                else:
                                        self.resultlist[funname]=[statement,branch,'NULL']
                            elif(int(statement)==int(self.resultlist[funname][0])):#statment部分等于
                                if(len(branch)>3):
                                    pass #branch项没内容，不操作
                                else:
                                    if(int(branch)>int(self.resultlist[funname][1])):#branch项大于
                                            self.resultlist[funname]=[statement,branch,'NULL']
                                    else:
                                        pass #小于等于，不操作
                            else:
                                pass  #statment比之前小，不进行任何操作
                        else:#未在字典中，添加
                            if(len(branch)>3):
                                    self.resultlist[funname]=[statement,'NULL','NULL']
                            else:
                                    self.resultlist[funname]=[statement,branch,'NULL']
                    else:
                        pass #statment等于0的那部分，意味着没测，不操作
                else:
                    print('第二次解析失败')
                    '''预留位置如出现其他状况需要解析，在此处增加'''
            else:#解析成功
                funname=r[0][0].replace("\n", "").replace("=20", "").replace("=", "").replace(" ", "")
                statement=r[0][1].replace("\n", "").replace("=20", "").replace("=", "").replace(" ", "")
                branch=r[0][2].replace("\n", "").replace("=20", "").replace("=", "").replace(" ", "")
                mcdc=r[0][3].replace("\n", "").replace("=20", "").replace("=", "").replace(" ", "")
                if(int(statement)>0):
                    if(funname in self.resultlist): #在字典中比较大小取最大
                        if(int(statement)>int(self.resultlist[funname][0])): #statment部分大于
                            if(len(branch)>3):
                                if(len(mcdc)>3):
                                    self.resultlist[funname]=[statement,'NULL','NULL']
                                else:
                                    self.resultlist[funname]=[statement,'NULL',mcdc]
                            else:
                                if(len(mcdc)>3):
                                    self.resultlist[funname]=[statement,branch,'NULL']
                                else:
                                    self.resultlist[funname]=[statement,branch,mcdc]
                        elif(int(statement)==int(self.resultlist[funname][0])):#statment部分等于
                            if(len(branch)>3):
                                pass #branch项没内容，不操作
                            else:
                                if(int(branch)>int(self.resultlist[funname][1])):#branch项大于
                                    if(len(mcdc)>3):
                                        self.resultlist[funname]=[statement,branch,'NULL']
                                    else:
                                        self.resultlist[funname]=[statement,branch,mcdc]
                                elif(int(branch)==int(self.resultlist[funname][1])):#branch项等于
                                        if(len(mcdc)>3):
                                            pass
                                        else:
                                            self.resultlist[funname]=[statement,branch,mcdc]
                                else:
                                    pass #小于，不操作
                        else:
                            pass  #statment比之前小，不进行任何操作
                    else:#未在字典中，添加
                        if(len(branch)>3):
                            if(len(mcdc)>3):
                                self.resultlist[funname]=[statement,'NULL','NULL']
                            else:
                                self.resultlist[funname]=[statement,'NULL',mcdc]
                        else:
                            if(len(mcdc)>3):
                                self.resultlist[funname]=[statement,branch,'NULL']
                            else:
                                self.resultlist[funname]=[statement,branch,mcdc]
                else:
                    pass #statment等于0的那部分，意味着没测，不操作

    #Deal LDRA TBrun Regression Report
    def dealRegression(self,text):
        caselist=re.findall(self.patternRows_1,text)
        casenum=str(len(caselist))
        tmptext=caselist[0]
        r=re.findall(self.patternfrom,tmptext)
        funname=r[0].replace("\n", "").replace("=20", "").replace("=", "").replace(" ", "")
        filename=r[1].replace("\n", "").replace("=20", "").replace("=", "").replace(" ", "")
        if(funname in self.casenumlist):
            if(int(casenum)>int(self.casenumlist[funname][0])):
                self.casenumlist[funname]=[casenum,filename]
            else:
                pass#案例数比原来小 不覆盖
        else:
            self.casenumlist[funname]=[casenum,filename]

    #操作Excel
    def makeExcel(self,resultlist,casenumlist,savepath):
        wb=openpyxl.Workbook()#创建Excel文件对象
        ws=wb.active #获取第一个sheet
        ws["A1"]='序号'
        ws["B1"]='文件名'
        ws["C1"]='函数名称'
        ws["D1"]='设计用例数'
        ws["E1"]='通过/未通过'
        ws["F1"]='语句覆盖率'
        ws["G1"]='分支覆盖率'
        ws["H1"]='MC/DC覆盖率'
        ws["I1"]='备注'
        '''设置列宽'''
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 21
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 11
        ws.column_dimensions['G'].width = 11
        ws.column_dimensions['H'].width = 11
        ws.column_dimensions['I'].width = 92
        '''设置单元格格式'''
        for nn in ['A','B','C','D','E','F','G','H','I']:
            ws[nn+'1'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
            ws[nn+'1'].font=Font(name='宋体', size=11,bold=True)
            border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin',color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin',color='000000'))
            ws[nn+'1'].border=border
        i=2
        for li in casenumlist:
            ws["A"+str(i)]=i-1
            ws["B"+str(i)]=casenumlist[li][1]
            ws["C"+str(i)]=li
            ws["D"+str(i)]=casenumlist[li][0]
            ws["E"+str(i)]='通过'
            if(li in resultlist):
                if(len(resultlist[li][0])>3):
                    ws["F"+str(i)]=resultlist[li][0]
                else:
                    ws["F"+str(i)]=resultlist[li][0]+"%"
                if(len(resultlist[li][1])>3):
                    ws["G"+str(i)]=resultlist[li][1]
                else:
                    ws["G"+str(i)]=resultlist[li][1]+"%"
                if(len(resultlist[li][2])>3):
                    ws["H"+str(i)]=resultlist[li][2]
                else:
                    ws["H"+str(i)]=resultlist[li][2]+"%"
            ws["I"+str(i)]=''

            '''设置单元格格式'''
            for nn in ['A','B','C','D','E','F','G','H','I']:
                ws[nn+str(i)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
                ws[nn+str(i)].font=Font(name='宋体', size=11)
                border = Border(left=Side(border_style='thin',color='000000'),
                                right=Side(border_style='thin',color='000000'),
                                top=Side(border_style='thin',color='000000'),
                                bottom=Side(border_style='thin',color='000000'))
                ws[nn+str(i)].border=border
            i+=1
        wb.save(savepath+"\\单元测试执行情况统计.xlsx")  #保存















if __name__ == '__main__':
    run=Deal()
    #递归遍历筛选文件
    def selectFile(dirpath):
        filelist=[]
        for root, dirs, files in os.walk(dirpath):
            #print(root) #当前目录路径
            #print(dirs) #当前路径下所有子目录
            #print(files) #当前路径下所有非目录子文件
            for file in files:
                if os.path.splitext(file)[1] == '.mht':
                    filelist.append(os.path.join(root, file))
        return filelist
    lis=selectFile('C:\\Users\\v5682\\Desktop\\新建\\1')
    for li in lis:
        print(li)
        text=run.getText(li)
        run.dealText(text)

    # print(len(run.resultlist))
    # for ii in run.resultlist:
    #     print(ii)
    #     print(run.resultlist[ii])

    # for oo in run.casenumlist:
    #     print(oo)
    #     print(run.casenumlist[oo])

    run.makeExcel(run.resultlist,run.casenumlist,'')