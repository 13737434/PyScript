import re
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill,colors
class Deal:
    #构建正则表达式
    patternistrue=re.compile('File\s*Contents\s*Report',re.S)#解析文本是否符合要求
    patterncontents=re.compile('File\s*Contents\s*Report\s*=*\n(.*?)\x0c',re.S)#解析Data Dictionary Summary Report页内容
    patternsome=re.compile(r'(\w+\.c)(.*?\n\n)',re.S) #解析.c文件名 +块
    patternCfiles=re.compile(r'Local\s*Functions(.*?)\n\n',re.S)#解析Global Functions块
    patternCfiles_1=re.compile(r'Global\s*Functions(.*?)\n\n',re.S)#解析文件名+Global Functions块
    patternfuns=re.compile('\w+',re.S)#解析每个.c文件所调用函数

    #获取 XXXXX.txt 中的文本
    def getString(self, url):
        with open(url, 'r') as f:
            text=f.read()
        return text

    # 判断文本是否符合要求
    def isTrue(self,text):
        items=re.findall(self.patternistrue,text)
        if(len(items)>0):
            return True
        else:
            return False

    #开始解析内容  .c文件名：【函数列表】
    def getCFileDict(self,text):
        fundic={}
        contents=re.findall(self.patterncontents,text)
        newtext=''.join(contents)+'\n'
        cfilelists=re.findall(self.patternsome,newtext)
        for cf in cfilelists:
            fromLF=re.findall(self.patternCfiles,cf[1])  #从Local Functions开始解析
            if(fromLF):#解析成功
                tmplists=re.findall(self.patternfuns,fromLF[0].replace("Global Functions", ""))
                fundic[cf[0]]=tmplists
            else:
                fromGF=re.findall(self.patternCfiles_1,cf[1]) #从Global Functions开始解析
                if(fromGF):
                    tmplists=re.findall(self.patternfuns,fromGF[0])
                    fundic[cf[0]]=tmplists
                else:
                    print(cf[0]+"中不存在函数")
        return fundic


    #Excel操作 全部
    def makeExcel_all(self,fundic,savepath):
        wb=openpyxl.Workbook()#创建Excel文件对象
        ws=wb.active #获取第一个sheet
        '''设置列宽'''
        ws.column_dimensions['A'].width =10 #序号
        ws.column_dimensions['B'].width = 35 #文件名
        ws.column_dimensions['C'].width =35 #函数名
        '''设置列标题'''
        ws["A1"]='序号'
        ws["B1"]='文件名'
        ws["C1"]='函数名'

        '''设置单元格格式'''
        for nn in ['A','B','C']:
            ws[nn+'1'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
            ws[nn+'1'].font=Font(name='宋体', size=11)
            border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin',color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin',color='000000'))
            ws[nn+'1'].border=border

        i=2
        for fd in fundic:
            for fd_1 in fundic[fd]:
                ws["A"+str(i)]=i-1
                ws["B"+str(i)]=fd
                ws["C"+str(i)]=fd_1
                '''设置单元格格式'''
                for nn in ['A','B','C']:
                    ws[nn+str(i)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
                    ws[nn+str(i)].font=Font(name='宋体', size=11)
                    border = Border(left=Side(border_style='thin',color='000000'),
                                right=Side(border_style='thin',color='000000'),
                                top=Side(border_style='thin',color='000000'),
                                bottom=Side(border_style='thin',color='000000'))
                    ws[nn+str(i)].border=border
                i+=1

        wb.save(savepath+'\\FunFrom.xlsx')


if __name__ == '__main__':
    dd=Deal()
    text=dd.getString('C:/Users/v5682/Desktop/MyUnderstandProject12.txt')
    if(dd.isTrue(text)):
        fundic=dd.getCFileDict(text)
        print("正在生成excel")
        dd.makeExcel_all(fundic,'C:/Users/v5682/Desktop')
        print("完成")
    else:
        print("文本不符合要求")