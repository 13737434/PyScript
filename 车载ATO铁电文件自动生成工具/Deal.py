import base64
import os
import chardet
from PyQt5.QtCore import Qt, pyqtSignal, QThread
from PyQt5.QtWidgets import QTableWidgetItem, QFileDialog, QMessageBox
from PyQt5 import QtGui, QtWidgets
from openpyxl import load_workbook
from 车载ATO铁电文件自动生成工具.photo import img
from 车载ATO铁电文件自动生成工具.MainShow import Ui_MainWindow
import sys
class MyPyQT_Form(QtWidgets.QMainWindow,Ui_MainWindow):
    signal=pyqtSignal(str) #定义信号，参数为str
    head=""
    table_1=[] #文本
    table_2=[] #excel
    def __init__(self):
        super(MyPyQT_Form,self).__init__()
        self.setupUi(self)
        '''UI Init Start'''
        tmp = open("phototmp233.jpg","wb+")
        tmp.write(base64.b64decode(img))
        tmp.close()
        self.lab_picture.setPixmap(QtGui.QPixmap("phototmp233.jpg")) #加载图片
        os.remove("phototmp233.jpg")
        #self.tableWidget_1.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) #表格列宽度均分
        self.headview_1=self.tableWidget_1.horizontalHeader()  #当前界面宽度 936
        self.headview_1.resizeSection(0,70)
        self.headview_1.resizeSection(1,270)
        self.headview_1.resizeSection(2,150)
        self.headview_1.resizeSection(3,150)

        self.headview_2=self.tableWidget_2.horizontalHeader()  #当前界面宽度 936
        self.headview_2.resizeSection(0,60)
        self.headview_2.resizeSection(1,270)
        self.headview_2.resizeSection(2,120)
        self.headview_2.resizeSection(3,70)
        self.headview_2.resizeSection(4,390)
        '''UI Init End'''
        '''Event Start'''
        self.btn_1.clicked.connect(self.addFile_1)
        self.btn_2.clicked.connect(self.addFile_2)
        self.btn_3.clicked.connect(self.makeFiles)
        '''Event End'''


    #添加通用ATO铁电参数文件
    def addFile_1(self):
        file, filetype = QFileDialog.getOpenFileName(self,"请选择通用ATO铁电参数文件","","Text Files (*.txt)")
        if(file== ""):
            self.textBrowser.append("取消选择1")
            return
        self.textBrowser.append("通用ATO铁电参数文件： "+file+"  导入成功")
        try:
            with open(file, 'rb') as f:
                f_read = f.read()
                f_charInfo = chardet.detect(f_read)
                text = f_read.decode(f_charInfo['encoding'])
        except:
            self.textBrowser.append("打开文件失败，请检查文件")
            return
        lines=text.split('\n')
        self.head=lines[0]
        tmptable=[]
        for i in range(1,len(lines)):
            cell=lines[i].split()
            if(len(cell)!=4):
                self.textBrowser.append("解析失败，请检查文本内容")
                return
            else:
                tmptable.append(cell)
        self.textBrowser.append("通用ATO铁电参数文件，解析成功")
        self.table_1=tmptable
        self.inputTable()

    #将解析后的通用ATO铁电参数加载到表格中
    def inputTable(self):
        self.tableWidget_1.setRowCount(0) #表格内容清除
        for i in range(0,len(self.table_1)):
            self.tableWidget_1.insertRow(i)
            self.tableWidget_1.setItem(i, 0, QTableWidgetItem(self.table_1[i][0]))
            self.tableWidget_1.setItem(i, 1, QTableWidgetItem(self.table_1[i][1]))
            self.tableWidget_1.setItem(i, 2, QTableWidgetItem(self.table_1[i][2]))
            self.tableWidget_1.setItem(i, 3, QTableWidgetItem(self.table_1[i][3]))
            self.tableWidget_1.item(i,0).setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter) #文本居中
            self.tableWidget_1.item(i,1).setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
            self.tableWidget_1.item(i,2).setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
            self.tableWidget_1.item(i,3).setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)

    #添加车组号与车辆参数对应关系表
    def addFile_2(self):
        file, filetype = QFileDialog.getOpenFileName(self,"请选择车组号与车辆参数对应关系表","","Excel Files (*.xlsx)")
        if(file== ""):
            self.textBrowser.append("取消选择2")
            return
        self.textBrowser.append("车组号与车辆参数对应关系表： "+file+"  导入成功")
        wb = load_workbook(file)
        wb.guess_types = True   #猜测格式类型
        ws=wb.active
        tmptable=[]
        if(ws.max_row==1):
            self.textBrowser.append("解析失败，请检查文件内容")
            return
        for i in range(1,ws.max_row+1):
            if(str(ws["A"+str(i)].value).isnumeric()):
                tmptable.append([str(ws["A"+str(i)].value),str(ws["B"+str(i)].value if ws["B"+str(i)].value!=None else ''),str(ws["C"+str(i)].value if ws["C"+str(i)].value!=None else ''),str(ws["D"+str(i)].value if ws["D"+str(i)].value!=None else ''),str(ws["E"+str(i)].value if ws["E"+str(i)].value!=None else '')])
            else:
                pass #表头内容不处理 或者其他情况
        self.table_2=tmptable
        self.textBrowser.append("车组号与车辆参数对应关系表，解析成功")
        self.inputTable_2()

    #将解析后的车组号与车辆参数对应关系表加载到表格中
    def inputTable_2(self):
        self.tableWidget_2.setRowCount(0) #表格内容清除
        for i in range(0,len(self.table_2)):
            self.tableWidget_2.insertRow(i)
            self.tableWidget_2.setItem(i, 0, QTableWidgetItem(self.table_2[i][0]))
            self.tableWidget_2.setItem(i, 1, QTableWidgetItem(self.table_2[i][1]))
            self.tableWidget_2.setItem(i, 2, QTableWidgetItem(self.table_2[i][2]))
            self.tableWidget_2.setItem(i, 3, QTableWidgetItem(self.table_2[i][3]))
            self.tableWidget_2.setItem(i, 4, QTableWidgetItem(self.table_2[i][4]))
            self.tableWidget_2.item(i,0).setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter) #文本居中
            self.tableWidget_2.item(i,1).setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
            self.tableWidget_2.item(i,2).setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
            self.tableWidget_2.item(i,3).setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
            self.tableWidget_2.item(i,4).setTextAlignment(Qt.AlignHCenter|Qt.AlignVCenter)

    #生成铁电参数文件
    def makeFiles(self):
        table_1=self.getTable_1()
        table_2=self.getTable_2()
        if(len(table_1)>0 and len(table_2)>0):  #两个表都准确无误 开始生成
            savepath=QFileDialog.getExistingDirectory(self, "请选择要生成的路径", "")
            if(savepath):
                # 创建线程
                self.thread = WriteThread(self.head,table_1,table_2,savepath)
                # 连接信号
                self.thread.signal.connect(self.receiveMess)
                # 开始线程
                self.thread.start()
            else:
                self.textBrowser.append("未选择生成路径")
        else:
            pass

    # 接收子线程消息   注意：0代表普通消息 1代表弹出消息
    def receiveMess(self,msg):
        if(msg[0]==0):
            self.textBrowser.append(msg[1])
        if(msg[0]==1):
            self.textBrowser.append(msg[1])
            QMessageBox.information(self,"Mess",msg[1])


    #获取通用ATO铁电参数
    def getTable_1(self):
        RowCont=self.tableWidget_1.rowCount()
        if(RowCont>0):
            table=[]
            for x in range(0,RowCont):
                tmplist=[]
                for y in range(0,4):
                    text=self.tableWidget_1.item(x, y).text()
                    if(text):
                        tmplist.append(text)
                    else:
                        self.textBrowser.append("通用ATO铁电参数文件中  第"+str(x+1)+"行,第"+str(y)+"列  存在空值")
                        return []
                table.append(tmplist)
            return table
        else:
            self.textBrowser.append("未导入通用ATO铁电参数文件")
            return []

    #获取车组号与车辆参数对应关系表
    def getTable_2(self):
        RowCont=self.tableWidget_2.rowCount()
        if(RowCont>0):
            table=[]
            for x in range(0,RowCont):
                tmplist=[]
                for y in range(0,4):
                    text=self.tableWidget_2.item(x, y).text()
                    if(text):
                        tmplist.append(text)
                    else:
                        self.textBrowser.append("车组号与车辆参数对应关系表中  第"+str(x+1)+"行,第"+str(y)+"列  存在空值")
                        return []
                table.append(tmplist)
            return table
        else:
            self.textBrowser.append("未导入车组号与车辆参数对应关系表文件")
            return []




#生成操作在该子线程中
class WriteThread(QThread):
    signal=pyqtSignal(list)
    def __init__(self,txthead,table_1,table_2,savepath):
        self.txthead=txthead
        self.table_1=table_1
        self.table_2=table_2
        self.savepath=savepath
        super(WriteThread, self).__init__()
    def run(self):
        self.make()

    def make(self):
        self.signal.emit([0,"开始生成。。。"])
        for li in self.table_2:
            tmptxt=self.table_1
            tmptxt[10][2]=li[3]
            tmptxt[10][3]=li[3]
            tmptxt[28][2]=li[2]
            tmptxt[28][3]=li[2]
            paraValue=[]
            for paraline in tmptxt:
                #值可能为16进制
                try:
                    paraValue_temp = int(paraline[2],10)
                except:
                    paraValue_temp = int(paraline[2],16)
                paraValue.append(paraValue_temp)
            data_crc = calculate_CRC32(paraValue,len(paraValue))
            #print ('data_crc',data_crc)
            data_crc_hex = hex(data_crc)
            #print('data_crc_hex',data_crc_hex )
            wcrc = '0x%x' %(data_crc)
            #print(wcrc)
            #line_crc_value_value = '10000'+'  '+'crc'+'  '+wcrc+'  '+wcrc
            #print 'line_crc_value_value',line_crc_value_value
            laststr=self.txthead+"\n".join(['   '.join(i)for i in tmptxt])+"\n; CRC_32\n10000   crc   "+wcrc+"   "+wcrc
            #print(laststr)
            try:
                self.writetxt(li[1],self.savepath,laststr)
                self.signal.emit([0,"序号："+li[0]+"   设备名称："+li[1]+"    生成成功。"])
            except Exception as e:
                self.signal.emit([1,"生成失败："+str(e)])
        self.signal.emit([1,"任务完成！！！"])

    #将字符串写入文本文件
    def writetxt(self,filename,savepath,text):
        with open(savepath+'/'+filename+'.txt', 'w',encoding='utf-8') as f:
            f.write(text)

'''外部导入'''
CRC16_XMODEM_TABLE = [
    0x0000, 0x1021, 0x2042, 0x3063, 0x4084, 0x50a5, 0x60c6, 0x70e7,
    0x8108, 0x9129, 0xa14a, 0xb16b, 0xc18c, 0xd1ad, 0xe1ce, 0xf1ef,
    0x1231, 0x0210, 0x3273, 0x2252, 0x52b5, 0x4294, 0x72f7, 0x62d6,
    0x9339, 0x8318, 0xb37b, 0xa35a, 0xd3bd, 0xc39c, 0xf3ff, 0xe3de,
    0x2462, 0x3443, 0x0420, 0x1401, 0x64e6, 0x74c7, 0x44a4, 0x5485,
    0xa56a, 0xb54b, 0x8528, 0x9509, 0xe5ee, 0xf5cf, 0xc5ac, 0xd58d,
    0x3653, 0x2672, 0x1611, 0x0630, 0x76d7, 0x66f6, 0x5695, 0x46b4,
    0xb75b, 0xa77a, 0x9719, 0x8738, 0xf7df, 0xe7fe, 0xd79d, 0xc7bc,
    0x48c4, 0x58e5, 0x6886, 0x78a7, 0x0840, 0x1861, 0x2802, 0x3823,
    0xc9cc, 0xd9ed, 0xe98e, 0xf9af, 0x8948, 0x9969, 0xa90a, 0xb92b,
    0x5af5, 0x4ad4, 0x7ab7, 0x6a96, 0x1a71, 0x0a50, 0x3a33, 0x2a12,
    0xdbfd, 0xcbdc, 0xfbbf, 0xeb9e, 0x9b79, 0x8b58, 0xbb3b, 0xab1a,
    0x6ca6, 0x7c87, 0x4ce4, 0x5cc5, 0x2c22, 0x3c03, 0x0c60, 0x1c41,
    0xedae, 0xfd8f, 0xcdec, 0xddcd, 0xad2a, 0xbd0b, 0x8d68, 0x9d49,
    0x7e97, 0x6eb6, 0x5ed5, 0x4ef4, 0x3e13, 0x2e32, 0x1e51, 0x0e70,
    0xff9f, 0xefbe, 0xdfdd, 0xcffc, 0xbf1b, 0xaf3a, 0x9f59, 0x8f78,
    0x9188, 0x81a9, 0xb1ca, 0xa1eb, 0xd10c, 0xc12d, 0xf14e, 0xe16f,
    0x1080, 0x00a1, 0x30c2, 0x20e3, 0x5004, 0x4025, 0x7046, 0x6067,
    0x83b9, 0x9398, 0xa3fb, 0xb3da, 0xc33d, 0xd31c, 0xe37f, 0xf35e,
    0x02b1, 0x1290, 0x22f3, 0x32d2, 0x4235, 0x5214, 0x6277, 0x7256,
    0xb5ea, 0xa5cb, 0x95a8, 0x8589, 0xf56e, 0xe54f, 0xd52c, 0xc50d,
    0x34e2, 0x24c3, 0x14a0, 0x0481, 0x7466, 0x6447, 0x5424, 0x4405,
    0xa7db, 0xb7fa, 0x8799, 0x97b8, 0xe75f, 0xf77e, 0xc71d, 0xd73c,
    0x26d3, 0x36f2, 0x0691, 0x16b0, 0x6657, 0x7676, 0x4615, 0x5634,
    0xd94c, 0xc96d, 0xf90e, 0xe92f, 0x99c8, 0x89e9, 0xb98a, 0xa9ab,
    0x5844, 0x4865, 0x7806, 0x6827, 0x18c0, 0x08e1, 0x3882, 0x28a3,
    0xcb7d, 0xdb5c, 0xeb3f, 0xfb1e, 0x8bf9, 0x9bd8, 0xabbb, 0xbb9a,
    0x4a75, 0x5a54, 0x6a37, 0x7a16, 0x0af1, 0x1ad0, 0x2ab3, 0x3a92,
    0xfd2e, 0xed0f, 0xdd6c, 0xcd4d, 0xbdaa, 0xad8b, 0x9de8, 0x8dc9,
    0x7c26, 0x6c07, 0x5c64, 0x4c45, 0x3ca2, 0x2c83, 0x1ce0, 0x0cc1,
    0xef1f, 0xff3e, 0xcf5d, 0xdf7c, 0xaf9b, 0xbfba, 0x8fd9, 0x9ff8,
    0x6e17, 0x7e36, 0x4e55, 0x5e74, 0x2e93, 0x3eb2, 0x0ed1, 0x1ef0,
]

CRC32_TABLE_LIST=[
    #��Ӧ������0xEDB88320(��0x04C11DB7�ľ���)
    0x00000000,0x77073096,0xee0e612c,0x990951ba,0x076dc419,0x706af48f,
    0xe963a535,0x9e6495a3,0x0edb8832,0x79dcb8a4,0xe0d5e91e,0x97d2d988,
    0x09b64c2b,0x7eb17cbd,0xe7b82d07,0x90bf1d91,0x1db71064,0x6ab020f2,
    0xf3b97148,0x84be41de,0x1adad47d,0x6ddde4eb,0xf4d4b551,0x83d385c7,
    0x136c9856,0x646ba8c0,0xfd62f97a,0x8a65c9ec,0x14015c4f,0x63066cd9,
    0xfa0f3d63,0x8d080df5,0x3b6e20c8,0x4c69105e,0xd56041e4,0xa2677172,
    0x3c03e4d1,0x4b04d447,0xd20d85fd,0xa50ab56b,0x35b5a8fa,0x42b2986c,
    0xdbbbc9d6,0xacbcf940,0x32d86ce3,0x45df5c75,0xdcd60dcf,0xabd13d59,
    0x26d930ac,0x51de003a,0xc8d75180,0xbfd06116,0x21b4f4b5,0x56b3c423,
    0xcfba9599,0xb8bda50f,0x2802b89e,0x5f058808,0xc60cd9b2,0xb10be924,
    0x2f6f7c87,0x58684c11,0xc1611dab,0xb6662d3d,0x76dc4190,0x01db7106,
    0x98d220bc,0xefd5102a,0x71b18589,0x06b6b51f,0x9fbfe4a5,0xe8b8d433,
    0x7807c9a2,0x0f00f934,0x9609a88e,0xe10e9818,0x7f6a0dbb,0x086d3d2d,
    0x91646c97,0xe6635c01,0x6b6b51f4,0x1c6c6162,0x856530d8,0xf262004e,
    0x6c0695ed,0x1b01a57b,0x8208f4c1,0xf50fc457,0x65b0d9c6,0x12b7e950,
    0x8bbeb8ea,0xfcb9887c,0x62dd1ddf,0x15da2d49,0x8cd37cf3,0xfbd44c65,
    0x4db26158,0x3ab551ce,0xa3bc0074,0xd4bb30e2,0x4adfa541,0x3dd895d7,
    0xa4d1c46d,0xd3d6f4fb,0x4369e96a,0x346ed9fc,0xad678846,0xda60b8d0,
    0x44042d73,0x33031de5,0xaa0a4c5f,0xdd0d7cc9,0x5005713c,0x270241aa,
    0xbe0b1010,0xc90c2086,0x5768b525,0x206f85b3,0xb966d409,0xce61e49f,
    0x5edef90e,0x29d9c998,0xb0d09822,0xc7d7a8b4,0x59b33d17,0x2eb40d81,
    0xb7bd5c3b,0xc0ba6cad,0xedb88320,0x9abfb3b6,0x03b6e20c,0x74b1d29a,
    0xead54739,0x9dd277af,0x04db2615,0x73dc1683,0xe3630b12,0x94643b84,
    0x0d6d6a3e,0x7a6a5aa8,0xe40ecf0b,0x9309ff9d,0x0a00ae27,0x7d079eb1,
    0xf00f9344,0x8708a3d2,0x1e01f268,0x6906c2fe,0xf762575d,0x806567cb,
    0x196c3671,0x6e6b06e7,0xfed41b76,0x89d32be0,0x10da7a5a,0x67dd4acc,
    0xf9b9df6f,0x8ebeeff9,0x17b7be43,0x60b08ed5,0xd6d6a3e8,0xa1d1937e,
    0x38d8c2c4,0x4fdff252,0xd1bb67f1,0xa6bc5767,0x3fb506dd,0x48b2364b,
    0xd80d2bda,0xaf0a1b4c,0x36034af6,0x41047a60,0xdf60efc3,0xa867df55,
    0x316e8eef,0x4669be79,0xcb61b38c,0xbc66831a,0x256fd2a0,0x5268e236,
    0xcc0c7795,0xbb0b4703,0x220216b9,0x5505262f,0xc5ba3bbe,0xb2bd0b28,
    0x2bb45a92,0x5cb36a04,0xc2d7ffa7,0xb5d0cf31,0x2cd99e8b,0x5bdeae1d,
    0x9b64c2b0,0xec63f226,0x756aa39c,0x026d930a,0x9c0906a9,0xeb0e363f,
    0x72076785,0x05005713,0x95bf4a82,0xe2b87a14,0x7bb12bae,0x0cb61b38,
    0x92d28e9b,0xe5d5be0d,0x7cdcefb7,0x0bdbdf21,0x86d3d2d4,0xf1d4e242,
    0x68ddb3f8,0x1fda836e,0x81be16cd,0xf6b9265b,0x6fb077e1,0x18b74777,
    0x88085ae6,0xff0f6a70,0x66063bca,0x11010b5c,0x8f659eff,0xf862ae69,
    0x616bffd3,0x166ccf45,0xa00ae278,0xd70dd2ee,0x4e048354,0x3903b3c2,
    0xa7672661,0xd06016f7,0x4969474d,0x3e6e77db,0xaed16a4a,0xd9d65adc,
    0x40df0b66,0x37d83bf0,0xa9bcae53,0xdebb9ec5,0x47b2cf7f,0x30b5ffe9,
    0xbdbdf21c,0xcabac28a,0x53b39330,0x24b4a3a6,0xbad03605,0xcdd70693,
    0x54de5729,0x23d967bf,0xb3667a2e,0xc4614ab8,0x5d681b02,0x2a6f2b94,
    0xb40bbe37,0xc30c8ea1,0x5a05df1b,0x2d02ef8d]

CRC32D_TABLE = [
    #/* CRC32生成多项式采用0x1EDC6F41 */
    0x00000000,0x1edc6f41,0x3db8de82,0x2364b1c3,0x7b71bd04,0x65add245,
    0x46c96386,0x58150cc7,0xf6e37a08,0xe83f1549,0xcb5ba48a,0xd587cbcb,
    0x8d92c70c,0x934ea84d,0xb02a198e,0xaef676cf,0xf31a9b51,0xedc6f410,
    0xcea245d3,0xd07e2a92,0x886b2655,0x96b74914,0xb5d3f8d7,0xab0f9796,
    0x05f9e159,0x1b258e18,0x38413fdb,0x269d509a,0x7e885c5d,0x6054331c,
    0x433082df,0x5deced9e,0xf8e959e3,0xe63536a2,0xc5518761,0xdb8de820,
    0x8398e4e7,0x9d448ba6,0xbe203a65,0xa0fc5524,0x0e0a23eb,0x10d64caa,
    0x33b2fd69,0x2d6e9228,0x757b9eef,0x6ba7f1ae,0x48c3406d,0x561f2f2c,
    0x0bf3c2b2,0x152fadf3,0x364b1c30,0x28977371,0x70827fb6,0x6e5e10f7,
    0x4d3aa134,0x53e6ce75,0xfd10b8ba,0xe3ccd7fb,0xc0a86638,0xde740979,
    0x866105be,0x98bd6aff,0xbbd9db3c,0xa505b47d,0xef0edc87,0xf1d2b3c6,
    0xd2b60205,0xcc6a6d44,0x947f6183,0x8aa30ec2,0xa9c7bf01,0xb71bd040,
    0x19eda68f,0x0731c9ce,0x2455780d,0x3a89174c,0x629c1b8b,0x7c4074ca,
    0x5f24c509,0x41f8aa48,0x1c1447d6,0x02c82897,0x21ac9954,0x3f70f615,
    0x6765fad2,0x79b99593,0x5add2450,0x44014b11,0xeaf73dde,0xf42b529f,
    0xd74fe35c,0xc9938c1d,0x918680da,0x8f5aef9b,0xac3e5e58,0xb2e23119,
    0x17e78564,0x093bea25,0x2a5f5be6,0x348334a7,0x6c963860,0x724a5721,
    0x512ee6e2,0x4ff289a3,0xe104ff6c,0xffd8902d,0xdcbc21ee,0xc2604eaf,
    0x9a754268,0x84a92d29,0xa7cd9cea,0xb911f3ab,0xe4fd1e35,0xfa217174,
    0xd945c0b7,0xc799aff6,0x9f8ca331,0x8150cc70,0xa2347db3,0xbce812f2,
    0x121e643d,0x0cc20b7c,0x2fa6babf,0x317ad5fe,0x696fd939,0x77b3b678,
    0x54d707bb,0x4a0b68fa,0xc0c1d64f,0xde1db90e,0xfd7908cd,0xe3a5678c,
    0xbbb06b4b,0xa56c040a,0x8608b5c9,0x98d4da88,0x3622ac47,0x28fec306,
    0x0b9a72c5,0x15461d84,0x4d531143,0x538f7e02,0x70ebcfc1,0x6e37a080,
    0x33db4d1e,0x2d07225f,0x0e63939c,0x10bffcdd,0x48aaf01a,0x56769f5b,
    0x75122e98,0x6bce41d9,0xc5383716,0xdbe45857,0xf880e994,0xe65c86d5,
    0xbe498a12,0xa095e553,0x83f15490,0x9d2d3bd1,0x38288fac,0x26f4e0ed,
    0x0590512e,0x1b4c3e6f,0x435932a8,0x5d855de9,0x7ee1ec2a,0x603d836b,
    0xcecbf5a4,0xd0179ae5,0xf3732b26,0xedaf4467,0xb5ba48a0,0xab6627e1,
    0x88029622,0x96def963,0xcb3214fd,0xd5ee7bbc,0xf68aca7f,0xe856a53e,
    0xb043a9f9,0xae9fc6b8,0x8dfb777b,0x9327183a,0x3dd16ef5,0x230d01b4,
    0x0069b077,0x1eb5df36,0x46a0d3f1,0x587cbcb0,0x7b180d73,0x65c46232,
    0x2fcf0ac8,0x31136589,0x1277d44a,0x0cabbb0b,0x54beb7cc,0x4a62d88d,
    0x6906694e,0x77da060f,0xd92c70c0,0xc7f01f81,0xe494ae42,0xfa48c103,
    0xa25dcdc4,0xbc81a285,0x9fe51346,0x81397c07,0xdcd59199,0xc209fed8,
    0xe16d4f1b,0xffb1205a,0xa7a42c9d,0xb97843dc,0x9a1cf21f,0x84c09d5e,
    0x2a36eb91,0x34ea84d0,0x178e3513,0x09525a52,0x51475695,0x4f9b39d4,
    0x6cff8817,0x7223e756,0xd726532b,0xc9fa3c6a,0xea9e8da9,0xf442e2e8,
    0xac57ee2f,0xb28b816e,0x91ef30ad,0x8f335fec,0x21c52923,0x3f194662,
    0x1c7df7a1,0x02a198e0,0x5ab49427,0x4468fb66,0x670c4aa5,0x79d025e4,
    0x243cc87a,0x3ae0a73b,0x198416f8,0x075879b9,0x5f4d757e,0x41911a3f,
    0x62f5abfc,0x7c29c4bd,0xd2dfb272,0xcc03dd33,0xef676cf0,0xf1bb03b1,
    0xa9ae0f76,0xb7726037,0x9416d1f4,0x8acabeb5]

#00crcУ��
def _crc16(data, crc, table):
    """Calculate CRC16 using the given table.
    `data`      - data for calculating CRC, must be a string
    `crc`       - initial value
    `table`     - table for caclulating CRC (list of 256 integers)
    Return calculated value of CRC
    """
    for byte in data:
        crc = ((crc<<8)&0xff00) ^ table[((crc>>8)&0xff)^ord(byte)]
    return crc & 0xffff

def crc16xmodem(data, crc=0xffff):
    """Calculate CRC-CCITT (XModem) variant of CRC16.
    `data`      - data for calculating CRC, must be a string
    `crc`       - initial value = 0xffff
    Return calculated value of CRC
    """
    return _crc16(data, crc, CRC16_XMODEM_TABLE)

def crc32_simp(buff,size):
    if buff != None:
        crc = 0xFFFFFFFF
        for  i in range(0,size):
            crc = CRC32_TABLE_LIST[(crc ^ buff[i]) & 0xFF] ^ (crc >> 8)

        crc = crc^0xFFFFFFFF

    else:

        crc = 0xffffffff

    return crc

def calculate_CRC32(aData,aSize):
    #print 'aData,aSize',aData,aSize
    #print ('aSize',aSize)
    if aData != None:
        nAccum = 0xffffffff
        for i in range(0,aSize):
            #print 'aData',aData[i]
            for j in range(0,4):
                tempdata = (aData[i]>>(8*(3-j)))&0xff
                #print 'tempdata',tempdata
                index = (nAccum>>24) ^ tempdata
                #print 'index',index
                nAccum = CRC32D_TABLE[index] ^ (( nAccum << 8)&0xffffffff)
                #print 'calculate_CRC32 nAccum', nAccum
        #nAccum = nAccum^0xFFFFFFFF
        #print 'calculate_CRC32 nAccum', nAccum
    else:
        nAccum = 0xffffffff

    #/*- 返回计算的CRC值 */
    #print ('nAccum',nAccum)
    return nAccum

'''END'''



if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    my_pyqt_form = MyPyQT_Form()
    my_pyqt_form.show()
    sys.exit(app.exec_())