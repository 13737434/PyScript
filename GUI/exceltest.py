import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors
wb=openpyxl.Workbook() #生成一个工作簿
# Font
ws=wb.active
ws.title='Font'
# default 11pt,Calibri
italic24Font=Font(size=24,italic=True)  #斜体24
ws['B3'].font=italic24Font
ws['B3']='24pt Italic'
boldRedFont=Font(name='Times New Roman',bold=True,color=colors.RED)
ws['A1'].font=boldRedFont
ws['A1']='Bold Red Time New Roman'

# Formulas
ws=wb.create_sheet('Formula')
ws['A1']=200
ws['A2']=300
ws['A3']='=SUM(A1:A2)'

# Setting row height and column width
ws=wb.create_sheet('dimensions')
ws['A1']='Tall row'
ws.row_dimensions[1].height=70
ws['B2']='Wide column'
ws.column_dimensions['B'].width=20

# Merging cells
ws=wb.create_sheet('merged')
ws.merge_cells('A1:D3')
ws['A1']='Twelve cells merged together'
ws.merge_cells('C5:D5')
ws['C5']='TWO merged cells'

# Unmerging cells
ws=wb.copy_worksheet(wb['merged'])
ws.title='unmerged'
ws.unmerge_cells('A1:D3')
ws.unmerge_cells('C5:D5')

wb.save('style.xlsx')



