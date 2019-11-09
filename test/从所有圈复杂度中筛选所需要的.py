'''从Excel  中加载所测函数列表'''
from openpyxl import load_workbook

wb = load_workbook("C:/Users/v5682/Desktop/单元执行情况/单元测试执行情况统计.xlsx")
wb.guess_types = True   #猜测格式类型
ws=wb.active
tmplist=[]
for i in range(2,450):
    if(ws["C"+str(i)].value):
        tmplist.append(ws["C"+str(i)].value)
print(len(tmplist))
print(tmplist)


'''加载静态度量的excel'''
wb1 = load_workbook("C:/Users/v5682/Desktop/静态度量问题报告单.xlsx")
wb1.guess_types = True   #猜测格式类型
ws1=wb1.active
tmplist1=[]
for i in range(12,121):
    if(ws1["C"+str(i)].value):
        tmplist1.append(ws["C"+str(i)].value)
print(len(tmplist1))
print(tmplist1)


for x in tmplist:
    for y in tmplist1:
        if(x==y):
            tmplist1.remove(x)
print(len(tmplist1))