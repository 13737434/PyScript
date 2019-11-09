from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill,colors

# '''统计.c文件'''
# tmp={}
# wb = load_workbook('C:/Users/v5682/Desktop/FunFrom.xlsx')
# wb.guess_types = True   #猜测格式类型
# ws=wb.active
# for i in range(2,1060):
#     tmp[ws["C"+str(i)].value]=ws["B"+str(i)].value
# print(len(tmp))
#
# '''统计调用关系'''
# tmp1={}
# wb1 = load_workbook('C:/Users/v5682/Desktop/函数调用关系(全部).xlsx')
# wb1.guess_types = True   #猜测格式类型
# ws1=wb1.active
# for i in range(2,981):
#     tmp1[ws1["C"+str(i)].value]=[ws1["B"+str(i)].value,ws1["D"+str(i)].value]
#
# print(len(tmp1))
#
#
#
# '''合并'''
# nonum=[]
# wb=Workbook()#创建Excel文件对象
# ws=wb.active #获取第一个sheet
# '''设置列宽'''
# ws.column_dimensions['A'].width =35
# ws.column_dimensions['B'].width = 35
# ws.column_dimensions['C'].width =35
# ws.column_dimensions['D'].width =35
# '''设置列标题'''
# ws["A1"]='文件名'
# ws["B1"]='入口函数名'
# ws["C1"]='调用level1'
# ws["D1"]='调用level2'
# '''设置单元格格式'''
# for nn in ['A','B','C','D']:
#     ws[nn+'1'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
#     ws[nn+'1'].font=Font(name='宋体', size=11)
#     border = Border(left=Side(border_style='thin',color='000000'),
#                     right=Side(border_style='thin',color='000000'),
#                     top=Side(border_style='thin',color='000000'),
#                     bottom=Side(border_style='thin',color='000000'))
#     ws[nn+'1'].border=border
# i=2
# for tt in tmp:
#     if tt in tmp1:
#         ws["A"+str(i)]=tmp[tt]
#         ws["B"+str(i)]=tmp1[tt][0]
#         ws["C"+str(i)]=tt
#         ws["D"+str(i)]=tmp1[tt][1]
#         '''设置单元格格式'''
#         for nn in ['A','B','C','D']:
#             ws[nn+str(i)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
#             ws[nn+str(i)].font=Font(name='宋体', size=11)
#             border = Border(left=Side(border_style='thin',color='000000'),
#                             right=Side(border_style='thin',color='000000'),
#                             top=Side(border_style='thin',color='000000'),
#                             bottom=Side(border_style='thin',color='000000'))
#             ws[nn+str(i)].border=border
#         i+=1
#     else:
#         print(tt+"不存在入口和调用")
#         nonum.append(tt)
#
# wb.save('C:/Users/v5682/Desktop/函数调用关系(.c文件).xlsx')
# print(len(nonum))


'''整理出.c文件名'''

ss=set()
wb = load_workbook('C:/Users/v5682/Desktop/FunFrom.xlsx')
wb.guess_types = True   #猜测格式类型
ws=wb.active
for i in range(2,1060):
    ss.add(ws["B"+str(i)].value)
print(ss)
print(len(ss))

wb=Workbook()#创建Excel文件对象
ws=wb.active #获取第一个sheet
i=1
for tt in ss:
    ws["A"+str(i)]=tt
    i+=1
wb.save('C:/Users/v5682/Desktop/tmp.xlsx')
