from openpyxl import load_workbook
##############################
####11111111
'''
fundic=dict()
wb = load_workbook("C:/Users/v5682/Desktop/单元测试执行情况统计.xlsx")
wb.guess_types = True   #猜测格式类型
ws=wb.active
for i in range(2,450):
    fundic[ws["C"+str(i)].value]=[ws["D"+str(i)].value,ws["F"+str(i)].value,ws["G"+str(i)].value,ws["H"+str(i)].value]
    #tmplist.append(ws["C"+str(i)].value)
print(len(fundic))
print(fundic)


wb1 = load_workbook("C:/Users/v5682/Desktop/FAO_atp_V0.0.1单元测试执行情况统计.xlsx")
wb1.guess_types = True   #猜测格式类型
for x in range(0,99):
    ws1=wb1.worksheets[x]
    for y in range(2,2000):
        if(ws1["C"+str(y)].value and ws1["C"+str(y)].value in fundic):
            ws1["D"+str(y)]=fundic[ws1["C"+str(y)].value][0]
            print(ws1["C"+str(y)].value)
            print(fundic[ws1["C"+str(y)].value][0])
            ws1["F"+str(y)]=fundic[ws1["C"+str(y)].value][1]
            ws1["G"+str(y)]=fundic[ws1["C"+str(y)].value][2]
            ws1["H"+str(y)]=fundic[ws1["C"+str(y)].value][3]
            #ws1["D"+str(i)]=fundic[]
wb1.save('C:/Users/v5682/Desktop/最终单元测试执行情况统计.xlsx')
'''
#######################

#222222
####
wb1 = load_workbook("C:/Users/v5682/Desktop/最终单元测试执行情况统计.xlsx")
wb1.guess_types = True   #猜测格式类型
for x in range(0,101):
    ws1=wb1.worksheets[x]
    for y in range(2,2000):
        if(str(ws1["F"+str(y)].value)=="100%"):
            ws1["I"+str(y)]=""
        if(str(ws1["G"+str(y)].value)=="null"):
            ws1["G"+str(y)]="NULL"
        if(str(ws1["H"+str(y)].value)=="null"):
            ws1["H"+str(y)]="NULL"
            #ws1["D"+str(i)]=fundic[]
wb1.save('C:/Users/v5682/Desktop/最终单元测试执行情况统计1.xlsx')