from openpyxl import load_workbook
ss=set()
wb = load_workbook('C:/Users/v5682/Desktop/ttt.xlsx')
wb.guess_types = True   #猜测格式类型
ws=wb.active
for i in range(1,148):
    ss.add(ws["A"+str(i)].value)
print(len(ss))
for i in ss:
    print(i)
