from openpyxl import load_workbook
dd={}
wb = load_workbook('C:/Users/v5682/Desktop/函数调用关系(全部).xlsx')
wb.guess_types = True   #猜测格式类型
ws=wb.active
for i in range(2,1678):
    dd[ws["C"+str(i)].value]=ws["D"+str(i)].value


wb1 = load_workbook('C:/Users/v5682/Desktop/tt.xlsx')
wb1.guess_types = True   #猜测格式类型
ws1=wb1.active
for i in range(2,240):
    if(ws1["D"+str(i)].value in dd):
        ws1["H"+str(i)]=dd[ws1["D"+str(i)].value]
    else:
        ws1["H"+str(i)]="--"
        print(ws1["D"+str(i)].value)

wb1.save('C:/Users/v5682/Desktop/tt1.xlsx')
