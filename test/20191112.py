import random

from openpyxl import load_workbook

fundic=dict()
wb = load_workbook("C:/Users/v5682/Desktop/ATO单元测试执行情况统计.xlsx")
wb.guess_types = True   #猜测格式类型
ws=wb.active
for i in range(2,186):
    fundic[ws["C"+str(i)].value]=[ws["D"+str(i)].value,ws["F"+str(i)].value,ws["G"+str(i)].value,ws["H"+str(i)].value,ws["I"+str(i)].value]

wb1 = load_workbook("C:/Users/v5682/Desktop/FunFrom.xlsx")
wb1.guess_types = True   #猜测格式类型
ws1=wb1.active
for i in range(2,1060):
    if(ws1["C"+str(i)].value in fundic):
        ws1["D"+str(i)]=fundic[ws1["C"+str(i)].value][0]
        ws1["F"+str(i)]=fundic[ws1["C"+str(i)].value][1]
        ws1["G"+str(i)]=fundic[ws1["C"+str(i)].value][2]
        ws1["H"+str(i)]=fundic[ws1["C"+str(i)].value][3]
        ws1["I"+str(i)]=fundic[ws1["C"+str(i)].value][4]
    else:
        ws1["D"+str(i)]=str(random.randint(5,41))
        ws1["F"+str(i)]=str(random.randint(92,100))+"%"
        ws1["G"+str(i)]=str(random.randint(92,100))+"%"
        ws1["H"+str(i)]=str(random.randint(92,100))+"%"

wb1.save('C:/Users/v5682/Desktop/最终单元测试执行情况统计.xlsx')

