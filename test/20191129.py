from openpyxl import load_workbook

list1=['zc_zession_init','zcs_login_logout_manage','zcs_login_logout_manage','zcs_train_locate','zcs_login_logout_manage','mod36_calculate_line_zc_sec_info','mod36_calculate_line_ats_sec_info',"appl_recur_inf_recv_balise_data",'appl_recur_inf_recv_RSSP_data_a','appl_recur_inf_recv_RSSP_data_b','app_inf_send_uart_log','cbtc_atp_appl_loop_run','atsse_ats_info_init','ats_session_init','atsse_run_out','atsse_set_ats_info',
       'headtail_comm_adap','cross_judge_with_ma','cross_line_run','cross_line_success','tbd_clear_next_line_id','ar_ctc_change_end_head']
wb = load_workbook("C:/Users/v5682/Desktop/函数调用关系.xlsx")
wb.guess_types = True   #猜测格式类型
ws=wb.active
for i in range(2,25):
    for li in list1:
        if(li in str(ws["D"+str(i)].value)):
             print(ws["C"+str(i)].value,"---------------",li)