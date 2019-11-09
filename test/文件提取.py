import shutil
import os
# 从excel第一列中读取所有.c文件名
from openpyxl import load_workbook

def getallfun(path):
    wb = load_workbook(path)
    wb.guess_types = True   #猜测格式类型
    ws=wb.active
    tmplist=[]
    for i in range(153,201):
        if(ws["B"+str(i)].value):
            tmplist.append(str(ws["B"+str(i)].value).replace(".c",""))
    print(len(tmplist))
    print(tmplist)
    return tmplist


#递归遍历筛选文件 把所需文件提取到一个文件夹
def selectFile(list,dirpath,dirpath_save):
    filelist=[]
    for root, dirs, files in os.walk(dirpath):
        #print(root) #当前目录路径
        #print(dirs) #当前路径下所有子目录
        #print(files) #当前路径下所有非目录子文件
        for file in files:
            if(os.path.splitext(file)[1] == '.c' and os.path.splitext(file)[0] in list):
                # filelist.append(os.path.join(root, file))
                list.remove(os.path.splitext(file)[0])
                shutil.copy(os.path.join(root, file), dirpath_save)
    if(list):
        print("没找到的文件")
        print(list)#未找到的文件名
    else:
        print("全部找到")

# 筛选所有头文件
def selectHeadFile(dirpath,dirpath_save):
    for root, dirs, files in os.walk(dirpath):
        #print(root) #当前目录路径
        #print(dirs) #当前路径下所有子目录
        #print(files) #当前路径下所有非目录子文件
        for file in files:
            if(os.path.splitext(file)[1] == '.h'):
                shutil.copy(os.path.join(root, file), dirpath_save)
    print("结束")

# 静态测试文件夹创建
def makedir(list,path):
    # 去除首位空格
    path=path.strip()
    # 去除尾部 \ 符号
    path=path.rstrip("\\")
    # 判断路径是否存在
    # 存在     True
    # 不存在   False
    isExists=os.path.exists(path)
    # 判断结果
    if not isExists:
        # 如果不存在则创建目录
        print(path+' 创建成功')
        # 创建目录操作函数
        os.makedirs(path)
        i=1
        for li in list:
            os.makedirs(path+"\\"+str(i)+"-"+li+".c")
            os.makedirs(path+"\\"+str(i)+"-"+li+".c"+"\\individual")
            os.makedirs(path+"\\"+str(i)+"-"+li+".c"+"\\system")
            i+=1
        return True
    else:
        # 如果目录存在则不创建，并提示目录已存在
        print(path+' 目录已存在')
        return False







if __name__ == '__main__':
    list=getallfun('C:/Users/v5682/Desktop/1.7.0白盒测试任务.xlsx')
    # makedir(list,"C:\\Users\\v5682\\Desktop\\static")
    selectFile(list,'C:\\Users\\v5682\\Desktop\\1.7.0代码','C:\\Users\\v5682\\Desktop\\新建\\董威')
    #selectHeadFile('C:\\Users\\v5682\\Desktop\\1.7.0代码','C:\\Users\\v5682\\Desktop\\新建\\head')