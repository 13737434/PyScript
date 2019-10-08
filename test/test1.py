#批量生成所需函数的函数调用关系




from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill,colors

se=set()
tmp=[]
list=[]
list2=[]
wb = load_workbook('C:/Users/v5682/Desktop/新建.xlsx')
wb.guess_types = True   #猜测格式类型
ws=wb.active
for i in range(1,449):
    se.add(ws["A"+str(i)].value)
    tmp.append(ws["A"+str(i)].value)
    list.append(ws["A"+str(i)].value)
print(len(se))
print(len(tmp))

#筛选出重复使用的函数
list_nohave=[]
for ss in se:
    tmp.remove(ss)
    list_nohave.append(ss)
print(len(tmp))
print(tmp)




wb = load_workbook('C:/Users/v5682/Desktop/函数调用关系(全部).xlsx')
wb.guess_types = True   #猜测格式类型
ws=wb.active

for i in range(1,1720):
    if(ws["C"+str(i)].value in se):
        list2.append([ws["B"+str(i)].value,ws["C"+str(i)].value,ws["D"+str(i)].value])
        list_nohave.remove(ws["C"+str(i)].value)
# print(list2)
print()
print(len(list_nohave))
print(list_nohave)
wb=Workbook()#创建Excel文件对象
ws=wb.active #获取第一个sheet
'''设置列宽'''
ws.column_dimensions['A'].width =10 #序号
ws.column_dimensions['B'].width = 35 #入口函数名
ws.column_dimensions['C'].width =35 #调用level1
ws.column_dimensions['D'].width =35 #调用level2
'''设置列标题'''
ws["A1"]='序号'
ws["B1"]='入口函数名'
ws["C1"]='调用level1'
ws["D1"]='调用level2'

'''设置单元格格式'''
for nn in ['A','B','C','D']:
    ws[nn+'1'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
    ws[nn+'1'].font=Font(name='宋体', size=11)
    border = Border(left=Side(border_style='thin',color='000000'),
                    right=Side(border_style='thin',color='000000'),
                    top=Side(border_style='thin',color='000000'),
                    bottom=Side(border_style='thin',color='000000'))
    ws[nn+'1'].border=border

i=2
for fd in list2:
    ws["A"+str(i)]=i-1
    ws["B"+str(i)]=fd[0]
    ws["C"+str(i)]=fd[1]
    ws["D"+str(i)]=fd[2]
    '''设置单元格格式'''
    for nn in ['A','B','C','D']:
        ws[nn+str(i)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) #水平、竖直对齐 自动换行
        ws[nn+str(i)].font=Font(name='宋体', size=11)
        border = Border(left=Side(border_style='thin',color='000000'),
                        right=Side(border_style='thin',color='000000'),
                        top=Side(border_style='thin',color='000000'),
                        bottom=Side(border_style='thin',color='000000'))
        ws[nn+str(i)].border=border
    i+=1

wb.save('C:/Users/v5682/Desktop/tmp.xlsx')