from openpyxl import load_workbook

text=''
with open('C:/Users/v5682/Desktop/test.txt', 'r',encoding='utf-8') as f:
    text=f.read()
print(text)


ttt=[]
wb = load_workbook('C:/Users/v5682/Desktop/HF3车载ID号码与ATO参数.xlsx')
wb.guess_types = True   #猜测格式类型
ws=wb.active
for i in range(3,93):
    ttt.append([ws["A"+str(i)].value,ws["B"+str(i)].value,ws["C"+str(i)].value,ws["D"+str(i)].value,ws["E"+str(i)].value])
    if(ws["E"+str(i)].value):
        m=text.replace("zhanwei2", str(ws["E"+str(i)].value)).replace("zhanwei3",str(ws["D"+str(i)].value ))
    else:
        m=text.replace("zhanwei2", str(0)).replace("zhanwei3",str(ws["D"+str(i)].value ))
    file = open('C:/Users/v5682/Desktop/新建文件夹/appl_parameter_conf_HF3-'+ws["B"+str(i)].value+'.txt', 'w',encoding='utf-8')
    file.write(m)

